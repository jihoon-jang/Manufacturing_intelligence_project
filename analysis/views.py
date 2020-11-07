from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.core.files.storage import FileSystemStorage

from django.db import connection

from analysis.models import CMAnalysis, CHUploadHistory, CHAnalysisStep, CMType, CHReport, CHAnalysisHistory, CHReportDetail, CHReportComment, CSMetaData, csmetadata2
import subprocess
import sys
import os.path
import datetime
import random
import requests

import openpyxl
from openpyxl import Workbook
from django.core import serializers
import json

from sqlalchemy import create_engine
DB_URL = 'mysql+mysqldb://root:jhw1996@223.194.46.212:3306/jejodb?charset=utf8'
engine = create_engine(DB_URL)

from django.views.decorators.csrf import csrf_exempt

from data_meta import *

# Create your views here.
def home(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    return render(request, "home.html", {
        'title':'Home',
        'menucd':'HOME',
        'member_id': member_id,
    })

def help(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    return render(request, "help.html", {
        'title':'Help',
        'menucd':'HOME',
        'member_id': member_id,
    })

def intelligence(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    return render(request, "intelligence.html", {
        'title':'Global Intelligence',
        'menucd':'INTELLIGENCE',
        'member_id': member_id,
    })
#--------------------------------------------------------------------------------------------------
# Analysis
#--------------------------------------------------------------------------------------------------
def analysis(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    if 'type_no' in request.GET:
        tno = int(request.GET['type_no'])
        rsAnalysis = CMAnalysis.objects.filter(usage_flag='1').filter(type_no=tno).all().order_by('-a_no')[:10]
    else:
        tno = 0
        rsAnalysis = CMAnalysis.objects.filter(usage_flag='1').all().order_by('type_no')[:10]

    rsType = CMType.objects.all().order_by('order_no')

    #print(rsType)

    return render(request, "analysis.html", {
        'rsAnalysis': rsAnalysis,
        'rsType':rsType,
        'title':'Analysis',
        'type_no':tno,
        'menucd':'ANALYSIS',
        'member_id': member_id,
    })

@csrf_exempt
def analysis_generate(request):

    a_datetime = datetime.datetime.today()
    a_title = '분석 ' + str(a_datetime)
    tno = request.GET['type_no']
    tdesc = request.GET['type_desc']
    print(tno)
    CMAnalysis.objects.create(a_title=a_title, a_note=a_title, a_owner=1, a_date=a_datetime, usage_flag='1', type_no=tno, type_desc=tdesc)

    response = redirect('/analysis')
    return response

# 파라메터 에러남...
def analysis_generate2(request):
    csr = connection.cursor()
    csr.callproc("p_cm_analysis_generate ", ())
    csr.close()

    response = redirect('/analysis')
    return response

def analysis_view(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    ano = request.GET['a_no']
    rows = CMAnalysis.objects.filter(a_no=ano)

    return render(request, "analysis_view.html", {
        'rsDetail': rows,
        'title':'View',
        'menucd': 'ANALYSIS',
        'member_id': member_id,
    })

def analysis_detail(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    ano = request.GET['a_no']
    rows = CMAnalysis.objects.filter(a_no=ano)

    return render(request, "analysis_detail.html", {
        'rsDetail': rows,
        'title':'Analysis > Run',
        'menucd': 'ANALYSIS',
        'member_id': member_id,
    })

def analysis_step(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    ano = int(request.GET['a_no'])
    row = CMAnalysis.objects.filter(a_no=ano).first()
    atitle = row.a_title

    rsStep = CHAnalysisStep.objects.filter(a_no=ano)

    return render(request, "analysis_step.html", {
        'rsStep': rsStep,
        'title':'분석단계',
        'a_no':ano,
        'a_title':atitle,
        'menucd': 'ANALYSIS',
        'member_id': member_id,
    })

def step_generate(request):

    ano = request.GET['a_no']
    atitle = request.GET['a_title']
    sdate = datetime.datetime.today()
    stitle = atitle + ' 단계 ' + str(sdate)

    CHAnalysisStep.objects.create(a_no=ano, step_title=stitle, upload_url='', upload_file='', step_date=sdate,
                                  usage_flag='1')

    analysis = CMAnalysis.objects.get(a_no=ano)
    analysis.analysis_cnt += 1
    analysis.save()

    resstr = '/analysis_step?a_no=' + ano
    return redirect(resstr)

@csrf_exempt
def analysis_run(request):
    ano = request.GET['a_no']
    sno = request.GET['step_no']
    hdate = datetime.datetime.today()

    data = {}

    afile = request.GET['upload_file']
    cmdstr = "import afiles." + afile
    # print(cmdstr)

    subprocess.call([sys.executable, "-c", cmdstr],)

    step = CHAnalysisStep.objects.get(step_no=sno)
    step.analysis_cnt += 1
    step.save()

    CHAnalysisHistory.objects.create(a_no=ano, step_no=sno, upload_url=afile, upload_file=afile,
                                   success_flag='Y',h_date=hdate)

    data['result_msg'] = '분석 CODE run success...'

    return JsonResponse(data, content_type="application/json")

@csrf_exempt
def analysis_update(request):
    ano = request.GET['a_no']
    atitle = request.GET['a_title']
    data = {}

    analysis = CMAnalysis.objects.get(a_no=ano)
    if atitle != "":
        analysis.a_title = atitle

    try:
        analysis.save()
        data['result_msg'] = '수정 하였습니다. updated...'
    except ValueError:
        data['result_msg'] = 'Update 에러. Error...'

    return JsonResponse(data, content_type="application/json")

@csrf_exempt
def analysis_toggle(request):
    ano = request.GET.get('a_no')

    analysis = CMAnalysis.objects.get(a_no=ano)
    data = {}

    analysis.usage_flag = '0'
    analysis.save()

    data['result_msg'] = '삭제 toggle 하였습니다. updated...'

    response = redirect('/analysis')
    return response


def analysis_toggle2(request):
    ano = request.GET['a_no']
    analysis = CMAnalysis.objects.get(a_no=ano)
    data = {}

    if analysis.usage_flag == '1':
        analysis.usage_flag = '0'
    else:
        analysis.usage_flag = '1'

    try:
        analysis.save()
        data['result_msg'] = '삭제 toggle 하였습니다. updated...'
    except ValueError:
        data['result_msg'] = '삭제 toggle. Error...'

    return JsonResponse(data, content_type="application/json")


def analysis_upload(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    ano = request.GET['a_no']
    sno = request.GET['step_no']

    context = {}
    context['a_no'] = ano
    context['step_no'] = sno
    context['title'] = 'Data processing'

    if CHAnalysisStep.objects.filter(step_no=sno).exists():
        rsStep = CHAnalysisStep.objects.get(step_no=sno)
        context['step_title'] = rsStep.step_title
        context['upload_file'] = rsStep.upload_file
        context['upload_url'] = rsStep.upload_url
        context['upload_cnt'] = rsStep.upload_cnt

        context['datatable_in'] = rsStep.datatable_in
        context['datatable_out'] = rsStep.datatable_out
        context['datain_cnt'] = rsStep.datain_cnt
        context['dataout_cnt'] = rsStep.dataout_cnt
        context['tablein_flag'] = rsStep.tablein_flag
        context['tableout_flag'] = rsStep.tableout_flag

        if rsStep.datatable_in == "1":
            context['rsDataIN'] = "in data"
        if rsStep.datatable_out == "1":
            context['rsDataOUT'] = "Out data"

    else:
        context['step_title'] = ""
        context['upload_file'] = ""
        context['upload_url'] = ""
        context['file_contents'] = ""
        context['datatable_in'] = "0"
        context['datatable_out'] = "0"
        context['datain_cnt'] = 0
        context['dataout_cnt'] = 0
        context['tablein_flag'] = "0"
        context['tableout_flag'] = "0"

    rsMetaIN = CSMetaData.objects.filter(step_no=sno, inout_flag ='IN')
    rsMetaOUT = CSMetaData.objects.filter(step_no=sno, inout_flag ='OUT')

    context['rsMetaIN'] = rsMetaIN
    context['rsMetaOUT'] = rsMetaOUT

    import pymysql
    dbCon = pymysql.connect(
        host = '127.0.0.1',
        port = 3307,
        user = 'root',
        passwd = 'jhw1996',
        db = 'jejodb')
    cursor = dbCon.cursor()

    cursor.execute("select * from tdata_" + str(sno) + "_in")
    rsDataIN = cursor.fetchall()

    cursor.execute("select * from tdata_" + str(sno) + "_out")
    rsDataOUT = cursor.fetchall()

    context['rsDataIN'] = rsDataIN
    context['rsDataOUT'] = rsDataOUT

    return render(request, "analysis_upload.html", context)

@csrf_exempt
def analysis_metawrite(request):

    sno = request.GET['step_no']
    context = {}
    idx = 1

    rsMetaIN = CSMetaData.objects.filter(step_no=sno, inout_flag='IN')
    rsMetaINJson = serializers.serialize('json', rsMetaIN)
    rsDataIN = json.loads(rsMetaINJson)

    # print(rsData)

    bookin = Workbook()
    sheet_in = bookin.active

    for i in rsDataIN:
        sheet_in.cell(row=1, column=idx).value = i['fields']['column_type']
        sheet_in.cell(row=2, column=idx).value = i['fields']['column_name']
        idx += 1

    filename_in = "static/datafiles/tdata_" + str(sno) + "_in.xlsx"
    bookin.save(filename_in)

    idx = 1

    rsMetaOUT = CSMetaData.objects.filter(step_no=sno, inout_flag='OUT')
    rsMetaOUTJson = serializers.serialize('json', rsMetaOUT)
    rsDataOUT = json.loads(rsMetaOUTJson)

    # print(rsData)

    bookout = Workbook()
    sheet_out = bookout.active

    for i in rsDataOUT:
        sheet_out.cell(row=1, column=idx).value = i['fields']['column_type']
        sheet_out.cell(row=2, column=idx).value = i['fields']['column_name']
        idx += 1

    filename_out = "static/datafiles/tdata_" + str(sno) + "_out.xlsx"
    bookout.save(filename_out)

    step = CHAnalysisStep.objects.get(step_no=sno)
    step.tablein_flag = "1"
    step.tableout_flag = "1"
    step.save()

    context['result_msg'] = 'Metadata written to \r\nINPUT : ' + filename_in + ' \r\nOUTPUT : ' + filename_out

    return JsonResponse(context, content_type="application/json")

@csrf_exempt
def analysis_metaread(request):

    sno = request.GET['step_no']
    type = request.GET['type']
    context = {}

    if type =="IN":
        filename = "static/datafiles/tdata_" + str(sno) + "_in.xlsx"
    else:
        filename = "static/datafiles/tdata_" + str(sno) + "_out.xlsx"

    book = openpyxl.load_workbook(filename)

    sheet = book.active
    max_col = sheet.max_column
    max_row = sheet.max_row

    rsData = []
    for j in range(1, max_row + 1):
        row = {}
        for i in range(1, max_col + 1):
            cell_title = sheet.cell(row=1, column=i).value
            cell_obj = sheet.cell(row=j, column=i)
            row[cell_title] = cell_obj.value
        rsData.append(row)

    # print(rsData)

    # for j in range(1, max_row + 1):
    #     print("========================")
    #     for i in range(1, max_col + 1):
    #         cell_obj = sheet.cell(row=j, column=i)
    #         print(cell_obj.value)

    context['rsData'] = rsData
    result_msg = 'Columns : ' + str(max_col) + ', Rows : ' + str(max_row)
    context['result_msg'] = result_msg

    return JsonResponse(context, content_type="application/json")

def analysis_metaupload(request):

    ano = request.POST['image_a_no']
    sno = request.POST['image_step_no']
    type = request.POST['image_type']

    data = {}

    if request.method == "POST":
        uploaded_file = request.FILES['ufile']
        name_old = uploaded_file.name
        name_ext = os.path.splitext(name_old)[1]
        name_new = 'tdata_' + str(sno) + '_' + type
        # print(name_new)

        file_name = name_new + name_ext
        #fs = FileSystemStorage()
        fs = FileSystemStorage(location='static/datafiles')
        if (fs.exists(file_name)):
            fs.delete(file_name)

        name = fs.save(file_name, uploaded_file)

        data['url'] = fs.url(name)
        data['result_msg'] = 'Data file uploaded...'

    else:
        data['result_msg'] = 'Data file upload error...'

    return redirect("/analysis_upload?a_no=" + ano + "&step_no=" + sno)



@csrf_exempt
def analysis_metasave(request):

    sno = request.GET['step_no']
    type = request.GET['type']
    context = {}

    if type == 'IN':
        table_name = 'tdata_' + str(sno) + '_in'
    else:
        table_name = 'tdata_' + str(sno) + '_out'

    rsMeta = CSMetaData.objects.filter(step_no=sno, inout_flag=type)

    strSQL = 'INSERT INTO ' + table_name + ' ('

    for i in rsMeta:
        strSQL += i.column_name + ','

    strSQL = strSQL[: -1] + ') VALUES ('

    #print(strSQL)

    if type == "IN":
        filename = "static/datafiles/" + table_name + ".xlsx"
    else:
        filename = "static/datafiles/" + table_name + ".xlsx"

    #print(filename)

    book = openpyxl.load_workbook(filename)

    sheet = book.active
    max_col = sheet.max_column
    max_row = sheet.max_row

    import pymysql
    dbCon = pymysql.connect(
        host = '127.0.0.1',
        port = 3307,
        user = 'root',
        passwd = 'jhw1996',
        db = 'jejodb')
    cursor = dbCon.cursor()

    if max_row > 2:
        for j in range(3, max_row + 1):
            strdata = ''
            for i in range(1, max_col + 1):
                if i == 1:
                    if sheet.cell(row=1, column=i).value == 'String':
                        strdata = "'" + str(sheet.cell(row=j, column=i).value) + "',"
                    else:
                        strdata = str(sheet.cell(row=j, column=i).value) + ","

                else:
                    if sheet.cell(row=1, column=i).value == 'String':
                        strdata += "'" + str(sheet.cell(row=j, column=i).value) + "',"
                    else:
                        strdata += str(sheet.cell(row=j, column=i).value) + ","

            strdata = strdata[:-1]
            strSQL2 = strSQL + strdata + ')'

            cursor.execute(strSQL2)

    dbCon.commit()

    if max_row > 2:
        data_cnt = max_row - 2
    else:
        data_cnt = 0

    if CHAnalysisStep.objects.filter(step_no=sno).exists():
        rsStep = CHAnalysisStep.objects.get(step_no=sno)

        if type =="IN":
            rsStep.datain_cnt = data_cnt
        else:
            rsStep.dataout_cnt = data_cnt

        rsStep.save()

    result_msg = 'Data saved... ' + type + ' : ' + str(data_cnt) + ' 건 uploaded'
    context['result_msg'] = result_msg

    return JsonResponse(context, content_type="application/json")


def analysis_code(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    ano = request.GET['a_no']
    sno = request.GET['step_no']

    context = {}
    context['a_no'] = ano
    context['step_no'] = sno
    context['title'] = '분석 Code'

    if CHAnalysisStep.objects.filter(step_no=sno).exists():
        rsStep = CHAnalysisStep.objects.get(step_no=sno)
        context['step_title'] = rsStep.step_title
        context['upload_file'] = rsStep.upload_file
        context['upload_url'] = rsStep.upload_url
        context['datain_cnt'] = rsStep.datain_cnt
        context['dataout_cnt'] = rsStep.dataout_cnt
        context['upload_cnt'] = rsStep.upload_cnt

        #print(rsStep.upload_file)

        if rsStep.upload_file == "":
            context['file_contents'] = ""
        else:
            file1 = "afiles/" + rsStep.upload_file + ".py"
            if os.path.isfile(file1) :

                f = open(file1, "r", encoding='utf8')

                context['file_contents'] = f.read()

    else:
        context['step_title'] = ""
        context['upload_file'] = ""
        context['upload_url'] = ""
        context['file_contents'] = ""
        context['datain_cnt'] = 0
        context['dataout_cnt'] = 0

    rsMetaIN = CSMetaData.objects.filter(step_no=sno, inout_flag ='IN')
    rsMetaOUT = CSMetaData.objects.filter(step_no=sno, inout_flag ='OUT')

    context['rsMetaIN'] = rsMetaIN
    context['rsMetaOUT'] = rsMetaOUT

    h_datetime = datetime.datetime.today()
    name_date = str(datetime.datetime.today().year) + '_' + str(datetime.datetime.today().month) + '_' + str(datetime.datetime.today().day)

    if request.method == "POST":
        uploaded_file = request.FILES['ufile']
        name_old = uploaded_file.name
        name_ext = os.path.splitext(name_old)[1]
        name_new = 'A' + name_date + '_' + str(random.randint(1000000000, 9999999999))

        fs = FileSystemStorage()
        #print(fs)

        name = fs.save(name_new + name_ext, uploaded_file)

        context['url'] = fs.url(name)
        context['member_id'] = member_id

        analysis = CHAnalysisStep.objects.get(step_no=sno)
        analysis.upload_url = fs.url(name)
        analysis.upload_file = name_new
        analysis.upload_cnt += 1
        analysis.save()

        CHUploadHistory.objects.create(a_no=ano, step_no=sno, upload_url=fs.url(name), upload_file=name_new, file_name=name_old, file_size=0, h_datetime=h_datetime)

        response = redirect('/analysis_step?a_no=' + ano)
        return response

    else:
        return render(request, "analysis_code.html", context)


@csrf_exempt
def analysisfile_update(request):

    sno = request.GET['step_no']
    fcontents = request.GET['file_contents']
    fcontents = fcontents.replace("<BR>", "\n")
    fcontents = fcontents.replace("<PLUS>", "+")
    fcontents = fcontents.replace("<SHARP>", "#python ")
    #print(fcontents)

    data = {}

    if CHAnalysisStep.objects.filter(step_no=sno).exists():
        rsStep = CHAnalysisStep.objects.get(step_no=sno)
        file1 = "afiles/" + rsStep.upload_file + ".py"

        f = open(file1, "w", encoding='utf8')
        f.write(fcontents)
        f.close()

    data['result_msg'] = "File updated..."

    return JsonResponse(data, content_type="application/json")

def analysis_data(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    ano = request.GET['a_no']
    sno = request.GET['step_no']
    context = {}

    if CHAnalysisStep.objects.filter(step_no=sno).exists():
        rsStep = CHAnalysisStep.objects.get(step_no=sno)
        stitle = rsStep.step_title
    else:
        stitle = "No title"

    # print(stitle)

    rsMetaIN = CSMetaData.objects.filter(step_no=sno, inout_flag ='IN')
    rsMetaOUT = CSMetaData.objects.filter(step_no=sno, inout_flag ='OUT')

    # print(rsMetaIN)

    return render(request, "analysis_data.html", {
        'rsMetaIN': rsMetaIN,
        'rsMetaOUT': rsMetaOUT,
        'title':'Metadata',
        'step_no':sno,
        'step_title':stitle,
        'a_no':ano,
        'menucd': 'ANALYSIS',
        'member_id': member_id,
    })


def analysis_sync_down(request):
    # https://stackoverflow.com/questions/9419162/download-returned-zip-file-from-url
    # https: // requests.readthedocs.io / en / master / user / quickstart /  # raw-response-content

    def download_url(url, save_path, chunk_size=128):
        r = requests.get(url, stream=True)
        with open(save_path, 'wb') as fd:
            for chunk in r.iter_content(chunk_size=chunk_size):
                fd.write(chunk)

    return "aaa"


@csrf_exempt
def step_update(request):
    sno = request.GET['step_no']
    stitle = request.GET['step_title']
    data = {}

    rsStep = CHAnalysisStep.objects.get(step_no=sno)
    if stitle != "":
        rsStep.step_title = stitle

    try:
        rsStep.save()
        data['result_msg'] = '단계 수정 updated...'
    except ValueError:
        data['result_msg'] = '단계 수정 Error...'

    return JsonResponse(data, content_type="application/json")


def data_column_generate(request):
    sno = request.GET['step_no']
    flag = request.GET['flag']
    cname = request.GET['column_name']
    ctype = request.GET['column_type']

    # print(sno)
    # print(flag)
    # print(cname)
    # print(ctype)
    cname = cname.strip()
    ctype = ctype.strip()

    data = {}

    with connection.cursor() as c1:
        strsql = "SELECT MAX(order_no) + 1 AS order_no FROM cs_meta_data WHERE step_no = " + str(sno)
        print(strsql)
        c1.execute(strsql)

        rows = c1.fetchone()

        o_no = rows[0]
        if o_no is None:
            o_no = 1

        #print(rows)

        conn = engine.connect()

        ins = csmetadata2.insert()
        conn.execute(ins, step_no=sno, inout_flag=flag, column_name=cname, column_type=ctype, order_no=o_no)

        conn.close()
        c1.close

        data['result_msg'] = 'Metadata column 생성 하였습니다.'

    return JsonResponse(data, content_type="application/json")

def data_columndelete(request):
    id = request.GET['id']

    data = {}

    if CSMetaData.objects.filter(id=id).exists():
        rs = CSMetaData.objects.get(id=id)
        rs.delete()

    data['result_msg'] = 'Column meta 삭제 하였습니다.'

    return JsonResponse(data, content_type="application/json")

def data_columncopy(request):
    sno = request.GET['step_no']

    data = {}

    strsql0 = "delete from cs_meta_data where step_no = " + str(sno) + " and inout_flag = 'OUT'"

    if CSMetaData.objects.filter(step_no=sno, inout_flag='OUT').exists():
        with connection.cursor() as c0:
            c0.execute(strsql0)
            rows = c0.fetchone()

    strsql = 'CALL p_cs_metadata_copy (' + str(sno) + ')'
    with connection.cursor() as c1:
        c1.execute(strsql)
        rows = c1.fetchone()

    c1.close()

    data['result_msg'] = 'Column meta Copy 하였습니다.'

    return JsonResponse(data, content_type="application/json")

def data_file_generate(request):
    sno = request.GET['step_no']
    flag = request.GET['flag']
    file_name = "data_meta.py"
    table_name = sno + "_" + flag
    data = {}

    #print(sno)
    print(flag)

    if os.path.exists(file_name):
        os.remove(file_name)

    #f = open(file_name, "x")

    f = open(file_name, "w+")
    f.write("from sqlalchemy.ext.declarative import declarative_base \n")
    f.write("Base = declarative_base() \n\n")
    f.write("from sqlalchemy import MetaData, Table, Column, Integer, String, ForeignKey, TIMESTAMP, DateTime, Float \n\n")
    f.write("metadata2 =  MetaData() \n\n")
    f.write("cnmeta_new = Table('tdata_" + table_name + "', metadata2, \n")
    f.write("               Column('id', Integer, primary_key=True),  \n")

    with connection.cursor() as c1:
        c1.execute("SELECT column_name, column_type FROM cs_meta_data WHERE step_no = %s and inout_flag = %s", { sno, flag })
        rows = c1.fetchall()

        print(rows)

        for x in rows:
            print(x[1])

            if x[1] == 'String':
                f.write("               Column('" + x[0] + "', " + x[1] + "(255)), \n")
            else:
                f.write("               Column('" + x[0] + "', " + x[1] + "), \n")

        c1.close

    f.write("               ) \n")

    f.close

    print("File write done... " + file_name)

    data['result_msg'] = 'Model file 파일 : ' + file_name + ", 테이블 : " + table_name

    return JsonResponse(data, content_type="application/json")

@csrf_exempt
def data_table_generate(request):

    sno = request.GET['step_no']
    dtype = request.GET['dtype']

    data = {}

    if dtype == "IN":
        strsql1 = "SHOW TABLES LIKE 'tdata_" + str(sno) + "_in'"
        strsql2 = "DROP TABLE tdata_" + str(sno) + "_in"
        strsql3 = "CREATE TABLE `tdata_" + str(sno) + "_in` ( \n"
        analysis = CHAnalysisStep.objects.get(step_no=sno)
        analysis.datatable_in = "1"
        analysis.save()

        fs = FileSystemStorage(location='static/datafiles')

        file_in = 'tdata_' + str(sno) + '_in'
        if (fs.exists(file_in)):
            fs.delete(file_in)

        #fs.closed

        analysis = CHAnalysisStep.objects.get(step_no=sno)
        analysis.tablein_flag = "0"
        analysis.save()

        rsMeta = CSMetaData.objects.filter(step_no=sno, inout_flag ='IN')

    if dtype == "OUT":
        strsql1 = "SHOW TABLES LIKE 'tdata_" + str(sno) + "_out'"
        strsql2 = "DROP TABLE tdata_" + str(sno) + "_out"
        strsql3 = "CREATE TABLE `tdata_" + str(sno) + "_out` ( \n"
        analysis = CHAnalysisStep.objects.get(step_no=sno)
        analysis.datatable_out = "1"
        analysis.save()

        fs = FileSystemStorage(location='static/datafiles')

        file_out = 'tdata_' + str(sno) + '_out'
        if (fs.exists(file_out)):
            fs.delete(file_out)

        #fs.closed

        analysis = CHAnalysisStep.objects.get(step_no=sno)
        analysis.tableout_flag = "0"
        analysis.save()

        rsMeta = CSMetaData.objects.filter(step_no=sno, inout_flag ='OUT')


    #print(rsMeta)
    #print(strsql1)
    with connection.cursor() as c1:
        c1.execute(strsql1)
        rows = c1.fetchone()

    c1.close()

    if rows:
        with connection.cursor() as c2:
            c2.execute(strsql2)
            rows = c2.fetchone()

        c2.close()

    strsql3 += "	`id` INT(11) NOT NULL AUTO_INCREMENT, \n"
    for i in rsMeta:
        if i.column_type == "String":
            strsql3 += "	`" + i.column_name.strip() + "` VARCHAR(255) NULL DEFAULT NULL, \n"
        elif i.column_type == "Integer":
            strsql3 += "	`" + i.column_name.strip() + "` INT(11) NULL DEFAULT NULL, \n"
        elif i.column_type == "Float":
            strsql3 += "	`" + i.column_name.strip() + "` FLOAT NULL DEFAULT NULL, \n"
        else:
            strsql3 += "	`" + i.column_name.strip() + "` VARCHAR(255) NULL DEFAULT NULL, \n"

    strsql3 += "	PRIMARY KEY (`id`) \n"
    strsql3 += ") \n"
    strsql3 += "COLLATE='utf8_general_ci' \n"
    strsql3 += "ENGINE=InnoDB \n"
    strsql3 += "; "

    #print(strsql3)
    with connection.cursor() as c3:
        c3.execute(strsql3)
        rows = c3.fetchone()

    c3.close()

    data['result_msg'] = 'Table created...'

    return JsonResponse(data, content_type="application/json")

#--------------------------------------------------------------------------------------------------
# Reporting
#--------------------------------------------------------------------------------------------------
@csrf_exempt
def analysis_report(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    ano = request.GET['a_no']
    rsAnalysis = CMAnalysis.objects.filter(usage_flag='1').all().order_by('-a_no')

    rsReport = CHReport.objects.filter(usage_flag='1').filter(a_no=ano).order_by('-r_no')[:10]

    return render(request, "analysis_report.html", {
        'rsReport': rsReport,
        'rsAnalysis':rsAnalysis,
        'title':'Reporting', 'a_no':ano,
        'menucd': 'ANALYSIS',
        'member_id': member_id,
    })

@csrf_exempt
def reporting(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    rsAnalysis = CMAnalysis.objects.filter(usage_flag='1').all().order_by('-a_no')

    rsReport = CHReport.objects.filter(usage_flag='1').filter(release_flag='1').order_by('-r_no')[:10]

    return render(request, "reporting.html", {
        'rsReport': rsReport,
        'rsAnalysis':rsAnalysis,
        'title':'Reporting',
        'menucd': 'REPORT',
        'member_id': member_id,
    })

def report_generate(request):

    ano = request.GET['a_no']
    rsAnalysis = CMAnalysis.objects.filter(a_no=ano).first()

    atitle = rsAnalysis.a_title
    rdate = datetime.datetime.today()
    rtitle = atitle + ' - Report ' + str(rdate)

    CHReport.objects.create(report_title=rtitle, a_no=ano, r_date=rdate, usage_flag='1', detail_cnt=0, char_cnt=0)

    rsAnalysis.reporting_cnt += 1
    rsAnalysis.reporting_date = rdate
    rsAnalysis.save()

    return redirect('/analysis_report?a_no=' + ano)


@csrf_exempt
def report_update(request):
    rno = request.GET['r_no']
    rtitle = request.GET['r_title']
    data = {}

    rsReport = CHReport.objects.get(r_no=rno)
    if rtitle != "":
        rsReport.report_title = rtitle

    try:
        rsReport.save()
        data['result_msg'] = 'Report 수정 updated...'
    except ValueError:
        data['result_msg'] = 'Report 수정 Error...'

    return JsonResponse(data, content_type="application/json")

def report_edit(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    rno = request.GET['r_no']
    ano = request.GET['a_no']

    row = CHReport.objects.filter(r_no=rno).first()
    report_title = row.report_title
    report_note = row.report_note
    report_user = row.member_id
    report_date = row.r_date
    detail_cnt = row.detail_cnt
    img_url = row.img_url
    release_flag = row.release_flag

    rsReport = CHReport.objects.filter(usage_flag='1').filter(a_no=ano).order_by('-r_no')[:10]

    rsDetail = CHReportDetail.objects.filter(r_no=rno).order_by('rh_no')[:100]

    return render(request, "report_edit.html", {
        'rsDetail': rsDetail,
        'rsReport': rsReport,
        'title':'Report Edit',
        'report_title':report_title,
        'report_note':report_note,
        'report_user':report_user,
        'report_date':report_date,
        'detail_cnt':detail_cnt,
        'img_url':img_url,
        'release_flag':release_flag,
        'a_no':ano,
        'r_no':rno,
        'menucd': 'ANALYSIS',
        'member_id': member_id,
    })

def report_view(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    rno = request.GET['r_no']

    row = CHReport.objects.filter(r_no=rno).first()
    ano = row.a_no
    report_title = row.report_title
    report_note = row.report_note
    report_user = row.member_id
    report_date = row.r_date
    detail_cnt = row.detail_cnt
    img_url = row.img_url
    release_flag = row.release_flag

    rsReport = CHReport.objects.filter(usage_flag='1').filter(a_no=ano).order_by('-r_no')[:10]

    rsDetail = CHReportDetail.objects.filter(r_no=rno).order_by('rh_no')[:100]

    rsComment = CHReportComment.objects.filter(r_no=rno).order_by('c_no')[:100]

    return render(request, "report_view.html", {
        'rsDetail': rsDetail,
        'rsReport': rsReport,
        'rsComment' : rsComment,
        'title':'Report View',
        'report_title':report_title,
        'report_note':report_note,
        'report_user':report_user,
        'report_date':report_date,
        'detail_cnt':detail_cnt,
        'img_url':img_url,
        'release_flag':release_flag,
        'a_no':ano,
        'r_no':rno,
        'menucd': 'REPORT',
        'member_id': member_id,
    })

@csrf_exempt
def reportcmt_create(request):

    rno = request.POST['image_r_no']
    nbox = request.POST['note_box']

    data = {}

    h_datetime = datetime.datetime.today()
    name_date = str(datetime.datetime.today().year) + '_' + str(datetime.datetime.today().month) + '_' + str(datetime.datetime.today().day)

    if request.method == "POST":
        if 'ufile' in request.FILES:
            print(1)
            uploaded_file = request.FILES['ufile']
            name_old = uploaded_file.name
            name_ext = os.path.splitext(name_old)[1]
            name_new = 'A' + name_date + '_' + str(random.randint(10000, 99999))

            #fs = FileSystemStorage()
            fs = FileSystemStorage(location='static/reportcmt')

            name = fs.save(name_new + name_ext, uploaded_file)

            imgurl = fs.url(name)
            imgname = name_new + name_ext

        else:
            imgurl = ""
            imgname = ""

        CHReportComment.objects.create(r_no=rno,
                                       comment_descript=nbox,
                                       image_url=imgurl,
                                       c_date=h_datetime)


        rs1 = CHReport.objects.get(r_no=rno)
        rs1.comment_cnt += 1
        rs1.save()

        data['result_msg'] = 'Image uploaded...'

    else:
        data['result_msg'] = 'Image upload error...'

    return redirect("/report_view?r_no=" + rno)

def reportedit_section(request):

    ano = request.GET['a_no']
    rno = request.GET['r_no']
    stitle = request.GET['s_title']

    CHReportDetail.objects.create(s_title=stitle, r_no=rno, s_note='(내용을 입력하세요)')

    return redirect('/report_edit?a_no=' + ano + '&r_no=' + rno)


@csrf_exempt
def reportedit_write(request):

    rno = request.GET['rno']
    rhno = request.GET['rhno']
    ctxt = request.GET['txt']
    flag = request.GET['flag']
    data = {}

    if flag == "title":
        detail = CHReportDetail.objects.get(rh_no=rhno)
        detail.s_title = ctxt
    elif flag == "pre":
        detail = CHReportDetail.objects.get(rh_no=rhno)
        detail.s_note = ctxt
    elif flag == "post":
        detail = CHReportDetail.objects.get(rh_no=rhno)
        detail.s_note_post = ctxt
    elif flag == "special":
        detail = CHReportDetail.objects.get(rh_no=rhno)
        detail.s_note_special = ctxt
    elif flag == "main":
        detail = CHReport.objects.get(r_no=rno)
        detail.report_note = ctxt

    try:
        detail.save()
        data['result_msg'] = 'Contents updated...'
    except ValueError:
        data['result_msg'] = 'Update error...'

    return JsonResponse(data, content_type="application/json")
    #return redirect("/report_edit?a_no=" + ano + "&r_no=" + rno)


def report_sectionimage_upload(request):
    member_id = None
    if request.session.has_key('member_id'):
        member_id = request.session['member_id']

    rhno = request.POST['image_rh_no']
    ano = request.POST['image_a_no']
    rno = request.POST['image_r_no']

    data = {}

    data['a_no'] = ano
    data['r_no'] = rno
    data['member_id'] = member_id

    h_datetime = datetime.datetime.today()
    name_date = str(datetime.datetime.today().year) + '_' + str(datetime.datetime.today().month) + '_' + str(datetime.datetime.today().day)

    if request.method == "POST":
        uploaded_file = request.FILES['ufile']
        name_old = uploaded_file.name
        name_ext = os.path.splitext(name_old)[1]
        name_new = 'A' + name_date + '_' + str(random.randint(1000000000, 9999999999))

        #fs = FileSystemStorage()
        fs = FileSystemStorage(location='static/photos')

        name = fs.save(name_new + name_ext, uploaded_file)

        data['url'] = fs.url(name)

        detail = CHReportDetail.objects.get(rh_no=rhno)
        #detail.img_url = fs.url(name)
        detail.img_url = name_new + name_ext
        detail.save()

        data['result_msg'] = 'Image uploaded...'

    else:
        data['result_msg'] = 'Image upload error...'

    return redirect("/report_edit?a_no=" + ano + "&r_no=" + rno)

@csrf_exempt
def reportedit_delete(request):
    rhno = request.GET['rhno']
    data = {}

    detail = CHReportDetail.objects.get(rh_no=rhno)

    try:
        detail.delete()
        data['result_msg'] = '삭제 되었습니다.'
    except ValueError:
        data['result_msg'] = '삭제 Error...'

    return JsonResponse(data, content_type="application/json")

@csrf_exempt
def report_release(request):
    rno = request.GET['rno']
    data = {}

    detail = CHReport.objects.get(r_no=rno)
    rflag = detail.release_flag

    if rflag == "1":
        detail.release_flag = "0"
    else:
        detail.release_flag = "1"

    try:
        detail.save()
        data['result_msg'] = '보고서 출시 토글...'
    except ValueError:
        data['result_msg'] = '출시 Error...'

    return JsonResponse(data, content_type="application/json")


#--------------------------------------------------------------------------------------------------
# Test
#--------------------------------------------------------------------------------------------------

#    subprocess.call(['python','script1.py'])
def subtest2(request):
    subprocess.call(['python.exe', 'dbtest2.py'])
    output = "<a href='/'>HOME</a><br><br><h1>subprocess test... dbtest2...,py runned</h1>"
    return HttpResponse(output, content_type='text/html')

# working : [sys.executable, "-c", "import django; print(django.get_version())"],

def subtest(request):
    cmdstr = "import afiles.A20200208-5841052667"
    subprocess.call(
        [sys.executable, "-c", cmdstr],
        )
    return redirect('/')

def subtest222(request):
    subprocess.call(
        [sys.executable, "-c", "import afiles.dbtest2_1"],
        [sys.executable, "-c", "import afiles.20200208_9215675035"],
        )
    return redirect('/')

def dbtest2(request):
    subprocess.call(['python.exe', 'dbtest2_1.py'])
    output = "<a href='/'>HOME</a><br><br><h1>subprocess test... dbtest2_1...,py runned</h1>"
    return HttpResponse(output, content_type='text/html')

def dbtest(request):
    subprocess.call(
        [sys.executable, "-c", "python dbtest2_1.py"],
        stdout=subprocess.PIPE,    # With nose process isolation, buffer can
        stderr=subprocess.STDOUT,  # easily get full and throw an error.
        )
    output = "<a href='/'>HOME</a><br><br><h1>subprocess test... dbtest2_1...,py runned</h1>"
    return HttpResponse(output, content_type='text/html')

    assert subprocess.call(
        [sys.executable, "-c", "import myproject"],
        cwd=self.tempdir,
        stdout=subprocess.PIPE,    # With nose process isolation, buffer can
        stderr=subprocess.STDOUT,  # easily get full and throw an error.
    ) == 0

def whatever(request):
    output = subprocess.run(['dir'], shell=True, capture_output=True, text=True, check=True)
    return HttpResponse(output, content_type='text/html')


from django.core.management.base import BaseCommand

from subprocess import Popen
from sys import stdout, stdin, stderr
import time, os, signal

class Command(BaseCommand):
    help = 'Run all commands'
    commands = [
        'redis-server',
        'python manage.py spider',
        'python manage.py schedule',
        'python manage.py postprocess',
        'python manage.py runserver'

    ]

    def handle(self, *args, **options):
        proc_list = []

        for command in self.commands:
            print
            "$ " + command
            proc = Popen(command, shell=True, stdin=stdin, stdout=stdout, stderr=stderr)
            proc_list.append(proc)

        try:
            while True:
                time.sleep(10)
        except KeyboardInterrupt:
            for proc in proc_list:
                os.kill(proc.pid, signal.SIGKILL)